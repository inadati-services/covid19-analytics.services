import os
import uuid
import datetime
from dateutil import parser
import settings as env
import openpyxl as px
from servant import file_downloader


def __convert_to_iso(date):
    # エクセルDatetimeをISOのDatetimeに変換する
    jst = datetime.timezone(datetime.timedelta(hours=+9), 'JST')
    str_date = date.isoformat(timespec='seconds')
    jst_iso_datetime = parser.parse(str_date).astimezone(jst)
    return jst_iso_datetime.isoformat()


def serve(db):
    # 既にエクセルファイルがある場合一旦、既存のファイルを破棄
    if os.path.isfile(env.DL_FILE_NAME):
        os.remove(env.DL_FILE_NAME)

    file_downloader.serve(env.DL_FILE, env.DL_FILE_NAME)

    # エクセルファイルを開いてシートを読みこむ
    wb = px.load_workbook(env.DL_FILE_NAME, data_only=True)
    sheet = wb["yousei"]

    date = __convert_to_iso(sheet.cell(row=sheet.max_row, column=1).value)

    db.execute("SELECT EXISTS (SELECT * FROM quantities WHERE date = %s);", (date,))
    (is_exist_on_db,) = db.fetchone()

    if not is_exist_on_db:
        _id = uuid.uuid4()
        prefecture_id = "f3344ece-9e94-488a-9811-59f852a4efe9"
        positive = sheet.cell(row=sheet.max_row, column=4).value - sheet.cell(row=sheet.max_row-1, column=4).value
        discharge = sheet.cell(row=sheet.max_row, column=9).value - sheet.cell(row=sheet.max_row-1, column=9).value
        death = sheet.cell(row=sheet.max_row, column=8).value - sheet.cell(row=sheet.max_row - 1, column=8).value
        sickbed_total = 509
        created_at = __convert_to_iso(datetime.datetime.now())

        db.execute("""
        INSERT
        INTO quantities (
            id,
            prefecture_id,
            date,
            positive,
            discharge,
            death,
            sickbed_total,
            created_at,
            updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
        """, (
            str(_id),
            prefecture_id,
            date,
            positive,
            discharge,
            death,
            sickbed_total,
            created_at,
            created_at
        ))
